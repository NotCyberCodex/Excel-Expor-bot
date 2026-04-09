import logging
import re
import os
from datetime import datetime
from telegram import Update
from telegram.ext import ApplicationBuilder, ContextTypes, CommandHandler, MessageHandler, filters
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
import dotenv

dotenv.load_dotenv()

# --- CONFIGURATION ---
GROUP_LINK = 'https://t.me/XZ_EARNING4658'
GROUP_USERNAME = '@XZ_EARNING4658'
CHANNEL_LINK = 'https://t.me/XZ_EARNING'
CHANNEL_USERNAME = '@XZ_EARNING'

# --- NEW: ADMIN CONFIGURATION ---
# Add your Telegram User ID here (integers only). Example: [123456789, 987654321]
ADMIN_IDS = [5390675752] 
USER_DB_FILE = "users.txt"

# --- LOGGING SETUP ---
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

# --- IN-MEMORY STORAGE ---
user_data_store = {}

# --- HELPER FUNCTIONS ---

def save_user_id(user_id):
    """
    Saves user ID to a file to track users for broadcasting.
    """
    if not os.path.exists(USER_DB_FILE):
        with open(USER_DB_FILE, "w") as f:
            pass
            
    with open(USER_DB_FILE, "r") as f:
        users = f.read().splitlines()
    
    if str(user_id) not in users:
        with open(USER_DB_FILE, "a") as f:
            f.write(f"{user_id}\n")

def get_all_users():
    """
    Returns a list of all user IDs from the file.
    """
    if not os.path.exists(USER_DB_FILE):
        return []
    with open(USER_DB_FILE, "r") as f:
        return [int(uid) for uid in f.read().splitlines() if uid.strip().isdigit()]

async def is_user_in_group(context, user_id):
    """
    Check if user is a member of the required group.
    """
    return True

def extract_c_user(cookie_string):
    match = re.search(r'c_user[:=]\s?["\']?(\d+)["\']?', cookie_string)
    if match:
        return match.group(1)
    return None

def is_duplicate_cookie(user_id, new_cookie):
    if user_id not in user_data_store or not user_data_store[user_id]['cookies']:
        return False
    
    new_uid = extract_c_user(new_cookie)
    if not new_uid:
        return False
        
    for existing_cookie in user_data_store[user_id]['cookies']:
        existing_uid = extract_c_user(existing_cookie)
        if existing_uid and existing_uid == new_uid:
            return True
    
    return False

# --- BOT COMMAND HANDLERS ---

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    # Save user ID for broadcasting
    save_user_id(user_id)
    
    user_data_store[user_id] = {'password': None, 'cookies': []}
    
    welcome_message = (
        "👋 Welcome!\n\n"
        "Please join our channel and group:\n"
        f"Channel: {CHANNEL_LINK}\n"
        f"Group: {GROUP_LINK}\n\n"
        "I convert Facebook cookies into an Excel file (Data Only).\n\n"
        "Steps:\n"
        "1. Set password: /setpassword <pass>\n"
        "2. Send your cookies (paste text directly)\n"
        "3. Download file: /export\n\n"
        "Type /clear to reset."
    )
    
    try:
        await update.message.reply_text(welcome_message, disable_web_page_preview=True)
    except Exception as e:
        try:
            await context.bot.send_message(chat_id=user_id, text=welcome_message, disable_web_page_preview=True)
        except Exception as e2:
            pass

async def set_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    # Save user ID just in case
    save_user_id(user_id)

    args = context.args

    if user_id not in user_data_store:
        user_data_store[user_id] = {'password': None, 'cookies': []}

    if not args:
        try:
            await update.message.reply_text("❌ Usage: /setpassword <your_password>", disable_web_page_preview=True)
        except Exception:
            try:
                await context.bot.send_message(chat_id=user_id, text="❌ Usage: /setpassword <your_password>", disable_web_page_preview=True)
            except Exception:
                pass
        return

    password = " ".join(args)
    user_data_store[user_id]['password'] = password
    try:
        await update.message.reply_text(f"✅ Password set to: {password}", disable_web_page_preview=True)
    except Exception:
        try:
            await context.bot.send_message(chat_id=user_id, text=f"✅ Password set to: {password}", disable_web_page_preview=True)
        except Exception:
            pass

async def handle_text_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    # Save user ID
    save_user_id(user_id)

    text = update.message.text

    if user_id not in user_data_store:
        user_data_store[user_id] = {'password': None, 'cookies': []}

    lines = text.split('\n')
    valid_count = 0
    duplicate_count = 0

    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        if extract_c_user(line):
            if is_duplicate_cookie(user_id, line):
                duplicate_count += 1
                continue
            
            user_data_store[user_id]['cookies'].append(line)
            valid_count += 1

    if valid_count > 0 or duplicate_count > 0:
        total = len(user_data_store[user_id]['cookies'])
        message = f"✅ Added {valid_count} cookies.\n📊 Total queued: {total}"
        if duplicate_count > 0:
            message += f"\n🔄 Skipped {duplicate_count} duplicate cookies."
        try:
            await update.message.reply_text(message, disable_web_page_preview=True)
        except Exception:
            try:
                await context.bot.send_message(chat_id=user_id, text=message, disable_web_page_preview=True)
            except Exception:
                pass
    else:
        try:
            await update.message.reply_text("⚠️ No valid cookies found (missing `c_user`).", disable_web_page_preview=True)
        except Exception:
            try:
                await context.bot.send_message(chat_id=user_id, text="⚠️ No valid cookies found (missing `c_user`).", disable_web_page_preview=True)
            except Exception:
                pass

async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    if user_id not in user_data_store:
        try:
            await update.message.reply_text("⚠️ No session active. Type /start.", disable_web_page_preview=True)
        except Exception:
            try:
                await context.bot.send_message(chat_id=user_id, text="⚠️ No session active. Type /start.", disable_web_page_preview=True)
            except Exception:
                pass
        return

    data = user_data_store[user_id]
    pwd = data.get('password')
    cookies = data.get('cookies')

    if not pwd:
        try:
            await update.message.reply_text("❌ Password not set! Use /setpassword <pass>", disable_web_page_preview=True)
        except Exception:
            try:
                await context.bot.send_message(chat_id=user_id, text="❌ Password not set! Use /setpassword <pass>", disable_web_page_preview=True)
            except Exception:
                pass
        return
    
    if not cookies:
        try:
            await update.message.reply_text("❌ No cookies found! Please send cookie strings first.", disable_web_page_preview=True)
        except Exception:
            try:
                await context.bot.send_message(chat_id=user_id, text="❌ No cookies found! Please send cookie strings first.", disable_web_page_preview=True)
            except Exception:
                pass
        return

    cookies_10 = []
    cookies_61 = []
    
    for cookie in cookies:
        uid = extract_c_user(cookie)
        if uid:
            if uid.startswith('10'):
                cookies_10.append(cookie)
            elif uid.startswith('61'):
                cookies_61.append(cookie)

    today = datetime.now().strftime("%m-%d-%Y")
    files_created = []
    
    if cookies_10:
        filename_10 = f"1000X  {today}.xlsx"
        if not create_excel_file(cookies_10, pwd, filename_10):
            try:
                await update.message.reply_text(f"❌ Failed to create Excel file '{filename_10}'.", disable_web_page_preview=True)
            except Exception:
                pass
            return
        files_created.append(filename_10)
    
    if cookies_61:
        filename_61 = f"61X  {today}.xlsx"
        if not create_excel_file(cookies_61, pwd, filename_61):
            try:
                await update.message.reply_text(f"❌ Failed to create Excel file '{filename_61}'.", disable_web_page_preview=True)
            except Exception:
                pass
            return
        files_created.append(filename_61)
    
    if files_created:
        for filename in files_created:
            try:
                with open(filename, 'rb') as file_doc:
                    await update.message.reply_document(document=file_doc, caption=f"✅ Here is your {filename} file.")
                os.remove(filename)
            except Exception:
                if os.path.exists(filename):
                    os.remove(filename)
                try:
                    await update.message.reply_text("❌ Failed to send file.", disable_web_page_preview=True)
                except Exception:
                    pass
    else:
        try:
            await update.message.reply_text("❌ No valid cookies found for export.", disable_web_page_preview=True)
        except Exception:
            pass
    
    user_data_store[user_id] = {'password': None, 'cookies': []}
    try:
        await update.message.reply_text("🧹 Data cleared.", disable_web_page_preview=True)
    except Exception:
        pass

def create_excel_file(cookies, password, filename):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Facebook Accounts"

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 100

        center_aligned = Alignment(horizontal='center', vertical='center')
        left_aligned = Alignment(horizontal='left', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for cookie in cookies:
            uid = extract_c_user(cookie)
            if uid:
                ws.append([uid, password, cookie])

        if ws.max_row >= 1:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                row[0].alignment = center_aligned
                row[0].border = thin_border
                row[1].alignment = center_aligned
                row[1].border = thin_border
                row[2].alignment = left_aligned
                row[2].border = thin_border

        wb.save(filename)
        return True
    except Exception as e:
        print(f"Error creating Excel file {filename}: {e}")
        return False

async def clear_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_data_store[user_id] = {'password': None, 'cookies': []}
    try:
        await update.message.reply_text("🗑️ Session data cleared.", disable_web_page_preview=True)
    except Exception:
        pass

# --- NEW: BROADCAST COMMAND ---
async def broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    # Check if user is an admin
    if user_id not in ADMIN_IDS:
        # Silent ignore or friendly error for non-admins
        return

    message = ' '.join(context.args)
    if not message:
        await update.message.reply_text("❌ Usage: /broadcast <your message here>")
        return

    users = get_all_users()
    if not users:
        await update.message.reply_text("❌ No users found in database.")
        return

    sent_count = 0
    blocked_count = 0
    
    status_msg = await update.message.reply_text(f"⏳ Broadcasting to {len(users)} users...")

    for uid in users:
        try:
            await context.bot.send_message(chat_id=uid, text=message, disable_web_page_preview=True)
            sent_count += 1
        except Exception as e:
            # User might have blocked the bot or deleted chat
            blocked_count += 1
            pass
    
    await context.bot.edit_message_text(
        chat_id=user_id,
        message_id=status_msg.message_id,
        text=f"✅ Broadcast Complete!\n\n📨 Sent: {sent_count}\n🚫 Failed/Blocked: {blocked_count}"
    )

# --- MAIN EXECUTION ---
if __name__ == '__main__':
    
    # NEON ASCII ART
    neon_art = """\033[1;36m
      __  __           _         _             _    _                       
     |  \/  |         | |       | |           | |  | |                      
     | \  / | __ _  __| | ___   | |__  _   _  | |__| | __ _ _ __ _ __ _   _ 
     | |\/| |/ _` |/ _` |/ _ \  | '_ \| | | | |  __  |/ _` | '__| '__| | | |
     | |  | | (_| | (_| |  __/  | |_) | |_| | | |  | | (_| | |  | |  | |_| |
     |_|  |_|\__,_|\__,_|\___|  |_.__/ \__, | |_|  |_|\__,_|_|  |_|   \__, |
                                        __/ |                          __/ |
                                       |___/                          |___/ \033[0m
    """
    print(neon_art)
    
    # ASK FOR TOKEN IN TERMINAL
    bot_token = input("\033[1;33mPlease enter your Telegram Bot Token: \033[0m").strip()
    
    if not bot_token:
        print("\033[1;31mError: No token provided. Exiting...\033[0m")
        exit(1)

    application = ApplicationBuilder().token(bot_token).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("setpassword", set_password))
    application.add_handler(CommandHandler("export", export_excel))
    application.add_handler(CommandHandler("clear", clear_data))
    
    # Register Broadcast Handler
    application.add_handler(CommandHandler("broadcast", broadcast))
    
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_message))

    print("\033[1;32mBot is running...\033[0m")

    application.run_polling()
